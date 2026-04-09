# -*- coding: utf-8 -*-

import pandas as pd
import numpy as np
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import warnings
warnings.filterwarnings('ignore')

# ─────────────────────────────────────────────────────────────────
# SEGMENT 1: READ DATA
# ─────────────────────────────────────────────────────────────────

FILE = 'Data.xlsx'

all_sheets    = pd.read_excel(FILE, sheet_name=None)
demand_df     = all_sheets['Demand_Data']
bom_df        = all_sheets['BOM']

demand_df.columns = demand_df.columns.str.strip()
bom_df.columns    = bom_df.columns.str.strip()

weeks         = demand_df['Week'].tolist()               # [1, 2, ..., 16]
forecast_base = demand_df['Forecast_Demand'].tolist()    # [50, 55, ..., 85]
NUM_WEEKS     = len(weeks)

print("Weeks        :", weeks)
print("Base Forecast:", forecast_base)
print(f"Total Weeks  : {NUM_WEEKS}")
print("\nBOM:\n", bom_df.to_string(index=False))

# SEGMENT 2: LOT SIZING RULE

def apply_lot_sizing(net_req, lot_size_rule):
    """
    L4L   -> order exactly the net requirement
    Fixed -> round up to the nearest multiple of the fixed lot size
    Examples:
        apply_lot_sizing(35, 'L4L') -> 35
        apply_lot_sizing(35, 50)    -> 50
        apply_lot_sizing(55, 50)    -> 100
    """
    if str(lot_size_rule).upper() == 'L4L':
        return net_req if net_req > 0 else 0
    else:
        lot = float(lot_size_rule)
        if net_req <= 0:
            return 0
        return np.ceil(net_req / lot) * lot

# SEGMENT 3: MRP ENGINE

def run_mrp(demand_series, bom_df, weeks, lot_policy='L4L'):
    """
    Runs full MRP explosion for all components in the BOM.

    Parameters
    ----------
    demand_series : list      weekly end-item (Bicycle) demand
    bom_df        : DataFrame BOM with columns:
                    Parent | Component | Level | Qty_Per | Lead_Time | Safety_Stock
    weeks         : list      week numbers [1, 2, ..., 16]
    lot_policy    : 'L4L' or a number (e.g. 50)
                    same rule applied uniformly to ALL components

    Returns
    -------
    dict  { component_name : { gross_req, proj_oh, planned_orders } }
    """

    results = {}
    n       = len(weeks)

    # Identify end item — the Parent at Level 1 is the top-level product
    parent_item = bom_df[bom_df['Level'] == 1]['Parent'].iloc[0]

    # Sort BOM top-down so every parent is processed before its children
    items = bom_df.sort_values('Level')

    for _, row in items.iterrows():

        component = row['Component']
        lead_time = int(row['Lead_Time'])
        ss        = float(row.get('Safety_Stock', 0))
        qty_per   = float(row['Qty_Per'])
        parent    = row['Parent']

        # lot_policy overrides whatever is in the BOM Lot_Size column
        lot_rule  = lot_policy

        # ── Gross Requirements ────────────────────────────────────────
        # Level-1 components -> driven by end-item demand x qty_per
        # Level-2 components -> driven by parent's planned orders x qty_per
        if parent == parent_item:
            gross_req = [demand_series[t] * qty_per for t in range(n)]
        else:
            parent_po = results.get(parent, {}).get('planned_orders', [0] * n)
            gross_req = [parent_po[t] * qty_per for t in range(n)]

        # ── Week-by-Week MRP Calculation ──────────────────────────────
        proj_oh        = [0.0] * n
        planned_orders = [0.0] * n

        for t in range(n):

            # On-hand carried from previous week (seed with safety stock at t=0)
            prev_oh = proj_oh[t - 1] if t > 0 else ss

            # Receive any planned order that was released (lead_time) weeks ago
            receipt = planned_orders[t - lead_time] if t >= lead_time else 0.0

            # Projected on-hand after fulfilling gross requirement this week
            proj_oh[t] = prev_oh + receipt - gross_req[t]

            # Net requirement: how much below safety stock are we?
            net_req = ss - proj_oh[t]

            if net_req > 0:
                po        = apply_lot_sizing(net_req, lot_rule)
                release_t = t - lead_time   # week to RELEASE the planned order

                if release_t >= 0:
                    planned_orders[release_t] += po
                    proj_oh[t]               += po   # update on-hand immediately
                else:
                    # Cannot push release further back than week 0
                    planned_orders[0] += po
                    proj_oh[t]        += po

        results[component] = {
            'gross_req':      gross_req,
            'proj_oh':        proj_oh,
            'planned_orders': planned_orders,
        }

    return results

# SEGMENT 4: NOISE LOGIC & ROLLING MRP (both policies)

def add_noise(base_forecast, noise_pct=0.10, seed=None):
    """
    Simulates real-world forecast revision by adding +-10% random noise.
    seed = run_week ensures each week's revision is different but reproducible
    so results can be reported consistently in the academic project.
    """
    rng = np.random.default_rng(seed)
    return [
        max(0, int(f * (1 + rng.uniform(-noise_pct, noise_pct))))
        for f in base_forecast
    ]


FIXED_LOT_SIZE = 50     # <- adjust to a realistic order quantity for your product

policies = {
    'L4L':       'L4L',
    'Fixed_Lot': FIXED_LOT_SIZE,
}

# Master store:
# { policy_name : { run_week : { component : { gross_req, proj_oh, planned_orders } } } }
all_rolling_results = {}

for policy_name, lot_policy in policies.items():

    rolling_results = {}

    for run_week in range(NUM_WEEKS):

        # Build revised forecast for this run:
        # Every week uses the noisy version of the base forecast.
        # No actual demand column — purely synthetic project.
        revised_forecast = add_noise(forecast_base, noise_pct=0.10, seed=run_week)

        rolling_results[run_week] = run_mrp(
            revised_forecast, bom_df, weeks, lot_policy=lot_policy
        )

    all_rolling_results[policy_name] = rolling_results
    print(f"Rolling MRP complete  ->  Policy: {policy_name}")

# SEGMENT 5: NERVOUSNESS METRICS


def compute_nervousness(rolling_results, bom_df, weeks):
    """
    Compares planned orders between every pair of consecutive MRP runs.

    A 'change' = any week where planned order quantity differs
                 between run[i] and run[i-1].

    Nervousness % = (total changes / total comparable periods) x 100

    Returns
    -------
    DataFrame : Component | Total_Periods | Changes | Nervousness_% | Stability_%
    """
    components = bom_df['Component'].unique()
    metrics    = []
    n          = len(weeks)

    for comp in components:

        changes = 0
        total   = 0

        # Collect planned-order list from each run in week order
        po_runs = [
            rolling_results[rw].get(comp, {}).get('planned_orders', [0] * n)
            for rw in sorted(rolling_results.keys())
        ]

        # Compare consecutive runs week by week
        for i in range(1, len(po_runs)):
            for t in range(n):
                prev   = po_runs[i - 1][t]
                curr   = po_runs[i][t]
                total += 1
                if abs(curr - prev) > 1e-6:
                    changes += 1

        nervousness_pct = round(100 * changes / total, 1) if total > 0 else 0.0

        metrics.append({
            'Component':     comp,
            'Total_Periods': total,
            'Changes':       changes,
            'Nervousness_%': nervousness_pct,
            'Stability_%':   round(100 - nervousness_pct, 1),
        })

    return pd.DataFrame(metrics)


# Compute nervousness for both policies
all_nervousness = {}
for policy_name, rolling_results in all_rolling_results.items():
    ndf           = compute_nervousness(rolling_results, bom_df, weeks)
    ndf['Policy'] = policy_name
    all_nervousness[policy_name] = ndf

# Combined comparison DataFrame
comparison_df = pd.concat(all_nervousness.values(), ignore_index=True)
comparison_df = comparison_df[['Policy', 'Component',
                                'Total_Periods', 'Changes',
                                'Nervousness_%', 'Stability_%']]

print("\n── Nervousness Policy Comparison ──────────────────────────")
print(comparison_df.to_string(index=False))

# SEGMENT 6: WRITE RESULTS TO EXCEL

wb = load_workbook(FILE)

# ── Reusable style definitions ────────────────────────────────────
HDR_FILL  = PatternFill('solid', start_color='1F4E79')   # dark blue  - main headers
HDR_FONT  = Font(bold=True, color='FFFFFF', name='Arial', size=10)
SUB_FILL  = PatternFill('solid', start_color='D6E4F0')   # light blue - sub-headers
SUB_FONT  = Font(bold=True, color='000000', name='Arial', size=10)
L4L_FILL  = PatternFill('solid', start_color='D6E4F0')   # blue tint  - L4L policy
FIX_FILL  = PatternFill('solid', start_color='FFE5B4')   # peach tint - Fixed policy
PO_FILL   = PatternFill('solid', start_color='E2EFDA')   # green tint - non-zero PO
RED_FILL  = PatternFill('solid', start_color='FF6B6B')   # red    - high nervousness
YEL_FILL  = PatternFill('solid', start_color='FFD93D')   # yellow - medium nervousness
GRN_FILL  = PatternFill('solid', start_color='6BCB77')   # green  - low nervousness
thin_side = Side(style='thin', color='BFBFBF')
BORDER    = Border(left=thin_side, right=thin_side,
                   top=thin_side,  bottom=thin_side)
CENTER    = Alignment(horizontal='center', vertical='center', wrap_text=True)
LEFT      = Alignment(horizontal='left',   vertical='center')


def style_header(cell, sub=False):
    cell.fill      = SUB_FILL if sub else HDR_FILL
    cell.font      = SUB_FONT if sub else HDR_FONT
    cell.alignment = CENTER
    cell.border    = BORDER


def style_body(cell, bold=False, fill=None, align='center'):
    cell.font      = Font(name='Arial', size=10, bold=bold)
    cell.alignment = CENTER if align == 'center' else LEFT
    cell.border    = BORDER
    if fill:
        cell.fill = fill


def set_width(ws, col_idx, width):
    ws.column_dimensions[get_column_letter(col_idx)].width = width


# ══════════════════════════════════════════════════════════════════
# SHEET 1: MRP_Summary
# Shows Gross Req, Projected On-Hand, Planned Orders for all
# components — final rolling run, both policies side by side.
# ══════════════════════════════════════════════════════════════════

if 'MRP_Summary' in wb.sheetnames:
    del wb['MRP_Summary']
ws1 = wb.create_sheet('MRP_Summary')

fixed_cols = ['Policy', 'Component', 'Metric']
all_cols   = fixed_cols + [f'W{w}' for w in weeks]

for c, h in enumerate(all_cols, 1):
    style_header(ws1.cell(1, c, h))

set_width(ws1, 1, 12)
set_width(ws1, 2, 22)
set_width(ws1, 3, 16)
for c in range(4, len(all_cols) + 1):
    set_width(ws1, c, 9)
ws1.row_dimensions[1].height = 22
ws1.freeze_panes = 'D2'

row = 2
for policy_name, rolling_results in all_rolling_results.items():
    last_run = rolling_results[NUM_WEEKS - 1]
    pol_fill = L4L_FILL if policy_name == 'L4L' else FIX_FILL

    for comp, data in last_run.items():
        for metric_name, values in [
            ('Gross Req',      data['gross_req']),
            ('Proj On-Hand',   data['proj_oh']),
            ('Planned Orders', data['planned_orders']),
        ]:
            is_po = (metric_name == 'Planned Orders')

            style_body(ws1.cell(row, 1, policy_name), bold=True,  fill=pol_fill)
            style_body(ws1.cell(row, 2, comp),        bold=is_po, align='left')
            style_body(ws1.cell(row, 3, metric_name), bold=is_po)

            for c, v in enumerate(values, 4):
                cell = ws1.cell(row, c, round(v, 1))
                style_body(cell, bold=is_po,
                           fill=PO_FILL if (is_po and v > 0) else None)
            row += 1
        row += 1    # blank spacer between components

    row += 1        # extra spacer between policies


# ══════════════════════════════════════════════════════════════════
# SHEET 2: Nervousness_Report
# Change counts and nervousness % for every component, both policies.
# Traffic-light colour coding on Nervousness_%.
# ══════════════════════════════════════════════════════════════════

if 'Nervousness_Report' in wb.sheetnames:
    del wb['Nervousness_Report']
ws2 = wb.create_sheet('Nervousness_Report')

for c, col_name in enumerate(comparison_df.columns, 1):
    style_header(ws2.cell(1, c, col_name))
    set_width(ws2, c, 18)
ws2.row_dimensions[1].height = 22

for r, row_data in enumerate(comparison_df.itertuples(index=False), 2):
    for c, val in enumerate(row_data, 1):
        cell = ws2.cell(r, c, val)

        if c == 1:      # Policy column
            fill = L4L_FILL if val == 'L4L' else FIX_FILL
            style_body(cell, bold=True, fill=fill)

        elif c == 5:    # Nervousness_% - traffic light
            if   val > 50: fill = RED_FILL
            elif val > 25: fill = YEL_FILL
            else:          fill = GRN_FILL
            style_body(cell, bold=True, fill=fill)

        elif c == 6:    # Stability_% - inverse traffic light
            if   val >= 75: fill = GRN_FILL
            elif val >= 50: fill = YEL_FILL
            else:           fill = RED_FILL
            style_body(cell, bold=True, fill=fill)

        else:
            style_body(cell)


# ══════════════════════════════════════════════════════════════════
# SHEET 3: Rolling_PO_Tracker
# Full matrix: rows = MRP runs, columns = weeks.
# One block per policy x component.
# Lets you visually trace how planned orders changed week-to-week.
# ══════════════════════════════════════════════════════════════════

if 'Rolling_PO_Tracker' in wb.sheetnames:
    del wb['Rolling_PO_Tracker']
ws3 = wb.create_sheet('Rolling_PO_Tracker')

components = bom_df['Component'].unique()
master_col = 1

for policy_name, rolling_results in all_rolling_results.items():
    pol_fill    = L4L_FILL if policy_name == 'L4L' else FIX_FILL
    block_width = NUM_WEEKS + 1   # run-label column + one column per week

    for comp in components:

        # Block title (merged across full width)
        title_cell           = ws3.cell(1, master_col,
                                        f'{policy_name}  |  {comp}  - Planned Orders')
        title_cell.fill      = HDR_FILL
        title_cell.font      = HDR_FONT
        title_cell.alignment = CENTER
        ws3.merge_cells(
            start_row=1, start_column=master_col,
            end_row=1,   end_column=master_col + block_width - 1
        )

        # Sub-header row
        style_header(ws3.cell(2, master_col, 'MRP Run / Week ->'), sub=True)
        set_width(ws3, master_col, 18)

        for w_idx, w in enumerate(weeks):
            style_header(ws3.cell(2, master_col + 1 + w_idx, f'W{w}'), sub=True)
            set_width(ws3, master_col + 1 + w_idx, 8)

        # Data rows — one per rolling MRP run
        for run_w in range(NUM_WEEKS):
            run_label = ws3.cell(3 + run_w, master_col, f'Run @ Wk {run_w + 1}')
            style_body(run_label, bold=True, fill=pol_fill)

            po_list = (rolling_results[run_w]
                       .get(comp, {})
                       .get('planned_orders', [0] * NUM_WEEKS))

            for w_idx, po in enumerate(po_list):
                cell = ws3.cell(3 + run_w, master_col + 1 + w_idx, round(po, 1))
                style_body(cell, fill=PO_FILL if po > 0 else None)

        master_col += block_width + 2   # gap between blocks


# ══════════════════════════════════════════════════════════════════
# SHEET 4: Policy_Comparison
# Clean side-by-side summary of L4L vs Fixed Lot nervousness.
# This is your key result table for the academic report.
# ══════════════════════════════════════════════════════════════════

if 'Policy_Comparison' in wb.sheetnames:
    del wb['Policy_Comparison']
ws4 = wb.create_sheet('Policy_Comparison')

# Section title
ws4.merge_cells('A1:F1')
title_cell           = ws4.cell(1, 1,
                                 'MRP Nervousness - L4L vs Fixed Lot Policy Comparison')
title_cell.fill      = HDR_FILL
title_cell.font      = Font(bold=True, color='FFFFFF', name='Arial', size=12)
title_cell.alignment = CENTER
ws4.row_dimensions[1].height = 28

# Column headers
for c, col_name in enumerate(comparison_df.columns, 1):
    style_header(ws4.cell(2, c, col_name))
    set_width(ws4, c, 18)
ws4.row_dimensions[2].height = 22

# Data rows
for r, row_data in enumerate(comparison_df.itertuples(index=False), 3):
    for c, val in enumerate(row_data, 1):
        cell = ws4.cell(r, c, val)

        if c == 1:      # Policy
            fill = L4L_FILL if val == 'L4L' else FIX_FILL
            style_body(cell, bold=True, fill=fill)

        elif c == 5:    # Nervousness_%
            if   val > 50: fill = RED_FILL
            elif val > 25: fill = YEL_FILL
            else:          fill = GRN_FILL
            style_body(cell, bold=True, fill=fill)

        elif c == 6:    # Stability_%
            if   val >= 75: fill = GRN_FILL
            elif val >= 50: fill = YEL_FILL
            else:           fill = RED_FILL
            style_body(cell, bold=True, fill=fill)

        else:
            style_body(cell)

# Legend block
leg_row = len(comparison_df) + 5
ws4.cell(leg_row, 1, 'LEGEND').font = Font(bold=True, name='Arial', size=10)

legend_items = [
    (leg_row + 1, GRN_FILL, 'Nervousness <= 25%  ->  Low  (Stable - Good)'),
    (leg_row + 2, YEL_FILL, 'Nervousness 26-50%  ->  Medium'),
    (leg_row + 3, RED_FILL, 'Nervousness > 50%   ->  High  (Unstable)'),
    (leg_row + 4, L4L_FILL, 'Policy: L4L - order exactly net requirement each week'),
    (leg_row + 5, FIX_FILL, f'Policy: Fixed Lot = {FIXED_LOT_SIZE} - '
                             f'round up to nearest {FIXED_LOT_SIZE}'),
]

for leg_r, leg_fill, leg_text in legend_items:
    colour_cell        = ws4.cell(leg_r, 1, '')
    colour_cell.fill   = leg_fill
    colour_cell.border = BORDER
    text_cell          = ws4.cell(leg_r, 2, leg_text)
    text_cell.font     = Font(name='Arial', size=9)
    text_cell.border   = BORDER
    ws4.merge_cells(start_row=leg_r, start_column=2,
                    end_row=leg_r,   end_column=6)


# ── Save workbook ─────────────────────────────────────────────────
wb.save(FILE)

print(f"\nAll results written to -> {FILE}")
print("  Sheets created:")
print("    MRP_Summary        - Gross Req, Proj OH, Planned Orders (both policies)")
print("    Nervousness_Report - Change counts per component")
print("    Rolling_PO_Tracker - Full rolling matrix across all runs")
print("    Policy_Comparison  - Side-by-side L4L vs Fixed Lot (key result)")

# ─────────────────────────────────────────────────────────────────
# GRAPHS SEGMENT — All 3 graphs matching your target images
# ─────────────────────────────────────────────────────────────────

import matplotlib
import matplotlib.pyplot as plt
import matplotlib.patches as mpatches
matplotlib.rcParams.update({
    'font.family':       'DejaVu Sans',
    'font.size':          11,
    'axes.titlesize':     13,
    'axes.titleweight':  'bold',
    'axes.spines.top':    False,
    'axes.spines.right':  False,
    'figure.dpi':         120,
})

BLUE       = '#1F4E79'
BLUE_LITE  = '#9DC3E6'
PEACH      = '#E07B39'
PEACH_LITE = '#F4B183'
GREY       = '#7F7F7F'

# ── Prep data ─────────────────────────────────────────────────────
l4l_data    = comparison_df[comparison_df['Policy'] == 'L4L'].reset_index(drop=True)
fixed_data  = comparison_df[comparison_df['Policy'] == 'Fixed_Lot'].reset_index(drop=True)
comp_labels = l4l_data['Component'].tolist()
x           = np.arange(len(comp_labels))
width       = 0.35

# ── Run WITHOUT SS rolling MRP  ───────────────
def build_rolling_forecast(base_forecast, run_week, num_weeks, noise_pct=0.15):
    """
    Realistic rolling forecast:
    Past weeks  (t < run_week) : frozen to base forecast
    Future weeks(t >= run_week): small noise applied
    Fixed seed ensures same noise pattern every run.
    """
    rng      = np.random.default_rng(42)
    forecast = []
    for t in range(num_weeks):
        if t < run_week:
            forecast.append(base_forecast[t])
        else:
            noisy = base_forecast[t] * (1 + rng.uniform(-noise_pct, noise_pct))
            forecast.append(max(0, int(noisy)))
    return forecast

bom_no_ss                 = bom_df.copy()
bom_no_ss['Safety_Stock'] = 0
rolling_no_ss             = {}

for policy_name, lot_policy in policies.items():
    roll_temp = {}
    for run_week in range(NUM_WEEKS):
        revised = build_rolling_forecast(forecast_base, run_week, NUM_WEEKS)
        roll_temp[run_week] = run_mrp(
            revised, bom_no_ss, weeks, lot_policy=lot_policy
        )
    rolling_no_ss[policy_name] = roll_temp
    print(f"No-SS rolling MRP complete -> Policy: {policy_name}")

# Also rerun WITH SS using build_rolling_forecast for fair comparison
rolling_with_ss = {}
for policy_name, lot_policy in policies.items():
    roll_temp = {}
    for run_week in range(NUM_WEEKS):
        revised = build_rolling_forecast(forecast_base, run_week, NUM_WEEKS)
        roll_temp[run_week] = run_mrp(
            revised, bom_df, weeks, lot_policy=lot_policy
        )
    rolling_with_ss[policy_name] = roll_temp

# ── Helper: per-run nervousness ───────────────────────────────────
def nervousness_per_run(rolling_results):
    n      = NUM_WEEKS
    values = []
    for i in range(1, NUM_WEEKS):
        changes = 0
        total   = 0
        for comp in bom_df['Component'].unique():
            po_prev = rolling_results[i-1].get(comp, {}).get('planned_orders', [0]*n)
            po_curr = rolling_results[i  ].get(comp, {}).get('planned_orders', [0]*n)
            for t in range(n):
                total += 1
                if abs(po_curr[t] - po_prev[t]) > 1e-6:
                    changes += 1
        values.append(round(100 * changes / total, 1) if total > 0 else 0.0)
    return values   # 15 values

# ── Helper: avg inventory per run ────────────────────────────────
def avg_inventory_per_run(rolling_results):
    values = []
    for rw in range(NUM_WEEKS):
        total_inv  = 0.0
        comp_count = 0
        for comp in bom_df['Component'].unique():
            oh = rolling_results[rw].get(comp, {}).get('proj_oh', [0]*NUM_WEEKS)
            total_inv  += sum(max(v, 0) for v in oh)
            comp_count += 1
        values.append(round(total_inv / comp_count, 1) if comp_count > 0 else 0.0)
    return values   # 16 values

# Compute all 4 combinations
nerv_l4l_with    = nervousness_per_run(rolling_with_ss['L4L'])
nerv_l4l_no      = nervousness_per_run(rolling_no_ss['L4L'])
nerv_fixed_with  = nervousness_per_run(rolling_with_ss['Fixed_Lot'])
nerv_fixed_no    = nervousness_per_run(rolling_no_ss['Fixed_Lot'])

inv_l4l_with     = avg_inventory_per_run(rolling_with_ss['L4L'])
inv_l4l_no       = avg_inventory_per_run(rolling_no_ss['L4L'])
inv_fixed_with   = avg_inventory_per_run(rolling_with_ss['Fixed_Lot'])
inv_fixed_no     = avg_inventory_per_run(rolling_no_ss['Fixed_Lot'])

run_numbers = list(range(1, NUM_WEEKS + 1))
run_labels  = [f'Wk{r}' for r in run_numbers]
run_pairs   = [f'R{i} vs R{i-1}' for i in range(1, NUM_WEEKS)]


# ══════════════════════════════════════════════════════════════════
# GRAPH 1: Policy Comparison for Nervousness
# ══════════════════════════════════════════════════════════════════

fig, ax = plt.subplots(figsize=(16, 6))

bars_l4l   = ax.bar(x - width/2, l4l_data['Nervousness_%'],
                    width, label='L4L',       color=BLUE,  alpha=0.88)
bars_fixed = ax.bar(x + width/2, fixed_data['Nervousness_%'],
                    width, label='Fixed Lot', color=PEACH, alpha=0.88)

for bar in bars_l4l:
    ax.text(bar.get_x() + bar.get_width() / 2,
            bar.get_height() + 0.8,
            f'{bar.get_height():.1f}%',
            ha='center', va='bottom',
            fontsize=8, color=BLUE, fontweight='bold')

for bar in bars_fixed:
    ax.text(bar.get_x() + bar.get_width() / 2,
            bar.get_height() + 0.8,
            f'{bar.get_height():.1f}%',
            ha='center', va='bottom',
            fontsize=8, color=PEACH, fontweight='bold')

ax.axhline(y=25, color=GREY, linestyle='--', linewidth=0.9, alpha=0.6)
ax.text(len(comp_labels) - 0.5, 26.5,
        'Low nervousness threshold (25%)', fontsize=8, color=GREY)

ax.set_xlabel('Component', labelpad=10)
ax.set_ylabel('Nervousness % (Changes)', labelpad=10)
ax.set_title('Graph 1: MRP Nervousness % by Component — L4L vs Fixed Lot Sizing Policy')
ax.set_xticks(x)
ax.set_xticklabels(comp_labels, rotation=40, ha='right', fontsize=9)
ax.legend(fontsize=11, loc='upper right')
ax.set_ylim(0, max(comparison_df['Nervousness_%'].max() + 20, 40))

plt.tight_layout()
plt.savefig('Policy_Comparison_for_Nervousness.png', dpi=150, bbox_inches='tight')
plt.show()
print("Graph 1 saved: Policy_Comparison_for_Nervousness.png")


# ══════════════════════════════════════════════════════════════════
# GRAPH 2: Planned Order Stability — Frame, Week 4 across all runs
# ══════════════════════════════════════════════════════════════════

TRACK_COMP = 'Frame'
TRACK_WEEK = 4
week_idx   = weeks.index(TRACK_WEEK)

l4l_po_track   = []
fixed_po_track = []

for run_w in range(NUM_WEEKS):
    l4l_po = (all_rolling_results['L4L'][run_w]
              .get(TRACK_COMP, {})
              .get('planned_orders', [0] * NUM_WEEKS))
    fixed_po = (all_rolling_results['Fixed_Lot'][run_w]
                .get(TRACK_COMP, {})
                .get('planned_orders', [0] * NUM_WEEKS))
    l4l_po_track.append(l4l_po[week_idx])
    fixed_po_track.append(fixed_po[week_idx])

fig, ax = plt.subplots(figsize=(14, 5))

ax.plot(range(NUM_WEEKS), l4l_po_track,
        marker='o', linewidth=2.5, markersize=7,
        color=BLUE,  linestyle='-', label='L4L')
ax.plot(range(NUM_WEEKS), fixed_po_track,
        marker='s', linewidth=2.5, markersize=7,
        color=PEACH, linestyle='-', label='Fixed Lot')

ax.fill_between(range(NUM_WEEKS), l4l_po_track,
                alpha=0.08, color=BLUE)
ax.fill_between(range(NUM_WEEKS), fixed_po_track,
                alpha=0.08, color=PEACH)

l4l_range   = max(l4l_po_track)   - min(l4l_po_track)
fixed_range = max(fixed_po_track) - min(fixed_po_track)

ax.text(0.02, 0.03,
        f'L4L variation range   : {l4l_range:.0f} units\n'
        f'Fixed Lot variation range: {fixed_range:.0f} units',
        transform=ax.transAxes, fontsize=9,
        verticalalignment='bottom',
        bbox=dict(boxstyle='round', facecolor='lightyellow', alpha=0.9))

ax.set_xlabel('MRP Run (Week Number)', labelpad=10)
ax.set_ylabel(f'Planned Order Qty — Week {TRACK_WEEK}', labelpad=10)
ax.set_title(f'Graph 2: Planned Order Stability — {TRACK_COMP} '
             f'(Week {TRACK_WEEK} tracked across all {NUM_WEEKS} runs)')
ax.set_xticks(range(NUM_WEEKS))
ax.set_xticklabels(run_labels, fontsize=9)
ax.legend(fontsize=11)

plt.tight_layout()
plt.savefig('SS_comparison_across_policies.png', dpi=150, bbox_inches='tight')
plt.show()
print("Graph 2 saved: SS_comparison_across_policies.png")


# ══════════════════════════════════════════════════════════════════
# GRAPH 3: Inventory — With SS vs Without SS, both policies
# ══════════════════════════════════════════════════════════════════

fig, ax = plt.subplots(figsize=(15, 6))

ax.plot(range(NUM_WEEKS), inv_l4l_no,
        marker='o', linewidth=2,   markersize=5,
        color=BLUE_LITE, linestyle='--',
        label='L4L — Without SS')
ax.plot(range(NUM_WEEKS), inv_l4l_with,
        marker='o', linewidth=2.5, markersize=7,
        color=BLUE,      linestyle='-',
        label='L4L — With SS')
ax.plot(range(NUM_WEEKS), inv_fixed_no,
        marker='s', linewidth=2,   markersize=5,
        color=PEACH_LITE, linestyle='--',
        label='Fixed Lot — Without SS')
ax.plot(range(NUM_WEEKS), inv_fixed_with,
        marker='s', linewidth=2.5, markersize=7,
        color=PEACH,      linestyle='-',
        label='Fixed Lot — With SS')

ax.fill_between(range(NUM_WEEKS),
                inv_l4l_no, inv_l4l_with,
                alpha=0.08, color=BLUE)
ax.fill_between(range(NUM_WEEKS),
                inv_fixed_no, inv_fixed_with,
                alpha=0.08, color=PEACH)

ax.set_xlabel('MRP Run (Week Number)', labelpad=10)
ax.set_ylabel('Avg On-Hand Inventory (units)', labelpad=10)
ax.set_title('Graph 3: Average Inventory — With vs Without Safety Stock\n'
             'L4L vs Fixed Lot Policy')
ax.set_xticks(range(NUM_WEEKS))
ax.set_xticklabels(run_labels, fontsize=9)
ax.legend(fontsize=9, ncol=2, loc='lower right')

ax.text(0.02, 0.03,
        'Solid lines  = With Safety Stock (higher inventory)\n'
        'Dashed lines = Without Safety Stock (lower inventory)',
        transform=ax.transAxes, fontsize=8,
        verticalalignment='bottom',
        bbox=dict(boxstyle='round', facecolor='lightyellow', alpha=0.9))

plt.tight_layout()
plt.savefig('Stability_of_policies_compared.png', dpi=150, bbox_inches='tight')
plt.show()
print("Graph 3 saved: Stability_of_policies_compared.png")

print("\nAll 3 graphs complete.")
print("  Policy_Comparison_for_Nervousness.png — Graph 1: bar chart per component")
print("  SS_comparison_across_policies.png     — Graph 2: planned order stability")
print("  Stability_of_policies_compared.png    — Graph 3: inventory with/without SS")